"""
    Convert an xls file into a csv with a given delimiter. The files will be called file1_sheet1name.csv,
    file2_sheet2name.csv, etc.

    Usage:

        python xls2csv.py delimiter file1 file2 ... filen

"""


from openpyxl import load_workbook
import sys
import os



def process(xls_file,delim):

    if not xls_file.endswith('.xlsx'):
        sys.stderr('Input should be an xlsx file')
    name_without_extension = xls_file[:-5]

    wb = load_workbook(xls_file)
    ws_all = wb.worksheets
    sheet_names = wb.sheetnames

    for ws_i, ws in enumerate(ws_all):

        sheet_file = name_without_extension+"_"+sheet_names[ws_i]+'.csv'
        data = [[i.value for i in j] for j in ws.rows]
        with open(sheet_file,'w') as out:
            for row in data:
                row2 = [str(ele) if ele is not None else "" for ele in row]
                line = delim.join(row2)+"\n"
                out.write(line.encode("utf-8"))


def main(delim,args):

    paths = []
    for arg in args:
        if os.path.isfile(arg):
            paths.append(arg)
        else:
            sys.stderr.write("  Error: unexpected argument `%s'\n" % arg)
            sys.exit()

    if not paths:
        sys.stderr.write("  Error: you must specify directories\n")
        sys.exit()

    for p in paths:
        process(p,delim)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
    else:
        main(sys.argv[1],sys.argv[2:])